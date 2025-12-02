import openpyxl

file_path = '[DNA] Forge Material Calculator (Shared).xlsx'

try:
    # Load the workbook, data_only=False ensures we get formulas
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    sheet_name = "Reset Time (UTC+7)"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Sheet: {sheet_name}")
        
        # Iterate through rows to find the 'Time Remaining' column and print formulas
        # Based on previous output, it looked like row 2 (index 1 in pandas) had data.
        # Let's just print the first 10 rows and columns A to F to see where the formulas are.
        
        for row in ws.iter_rows(min_row=1, max_row=10, max_col=6):
            row_data = []
            for cell in row:
                # cell.value will be the formula if it exists
                row_data.append(f"{cell.coordinate}: {cell.value}")
            print(" | ".join(row_data))
            
    else:
        print(f"Sheet '{sheet_name}' not found.")

except Exception as e:
    print(f"Error: {e}")
